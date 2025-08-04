from file_worker import read_json,add_line
import os
import pandas
from openpyxl import Workbook, load_workbook
import json
import tqdm
import pandas as pd
import re
import glob


def get_full_text(timestamps):
    text = ""
    for timestamp in timestamps:
        text+= timestamp["text"] + " "
    return text

# def extract_number(filename):
#     # Регулярное выражение для поиска числа между "audio_" и ".wav"
#     match = re.search(r'audio_(\d+)\.wav', filename)
#     if match:
#         return int(match.group(1))

#     match = re.search(r'audio_(\d+)\.json', filename)
#     if  match:
#         return int(match.group(1))
    
#     return None


def extract_number(filename):
    # # Регулярное выражение для поиска числа между "audio_" и ".wav"
    # match = re.search(r'audio_(\d+)\.wav', filename)
    # if match:
    #     return int(match.group(1))
    # else:
    #     return None
    # Используем регулярное выражение для поиска первого числа
    match = re.search(r'\d+', filename)
    # Если число найдено, возвращаем его как целое число, иначе возвращаем None
    return int(match.group()) if match else None

    
def get_audios_paths(data_path):

    audios_paths = []
    print("data path",data_path,os.path.exists(data_path))
    for dirpath, dirnames, filenames in  os.walk(data_path):
        for filename in filenames:
            audios_paths.append((dirpath,filename))
    return sorted(audios_paths,key = lambda x:extract_number(x[1]))

def read_excel_to_df(file_path,sheet_name = 0):
    # Чтение Excel файла
    ending = file_path.split(".")[-1]
    if ending == "csv":
        return pd.read_csv(file_path)
    return pd.read_excel(file_path,sheet_name)
         

path = "/home/jovyan/Nikita/speech2latex/SberSaluth/whisper_synthesized_audios_final"
path_result = "/home/jovyan/Nikita/speech2latex/SberSaluth/whisper_synthesized_audios_final_excel"
path_output = path_result + "/dataset_match_3.xlsx"
speakers_file = path_result + "/speakers_3.txt"
error_log_path = "./errors.txt"

workbook = Workbook()
sheet = workbook.active
sheet.title = "truly_checked_new_formulas"
headers = ["speaker_id","language","formula_id","is_tts","pronunciation","whisper_transcription","latex","audio_path"]

sheet.append(headers)
row = sheet.max_row + 1 if sheet.max_row > 1 else 1

speaker_names = []


pbar = tqdm.tqdm(glob.glob(f"{path}/**/*.json",recursive=True))

for filename in pbar:
    pbar.set_description(f"{filename}")
    try:
        speaker_path = os.path.dirname(filename)
        speaker_name = os.path.basename(speaker_path)

        if speaker_name not in speaker_names:
            speaker_names.append(speaker_name)

        speaker_id = speaker_names.index(speaker_name)
        
        # filename = os.path.join(speaker_path,file)
        row_dict = read_json(filename)

        # words_timestamps = json.dumps(row_dict["words_timestamps"])
        sentences_timestamps = get_full_text(row_dict["sentences_timestamps"])
        file_name = str(row_dict["file_name"])
        pronunciation = row_dict["pronunciation"]
        latex = row_dict["latex"]
        audio_path = row_dict["audio_path"]


        is_tts = 1 if os.path.basename(os.path.dirname(os.path.dirname(speaker_path))) == "tts" else 0 
        language =  os.path.basename(os.path.dirname(speaker_path))

        data_row = [speaker_id,language,file_name,is_tts,pronunciation,sentences_timestamps,latex,audio_path]
        sheet.append(data_row)
    except Exception as ex:
        info = f"speaker_id {speaker_id} speaker_name {speaker_name} filename {filename} {ex}"
        add_line(error_log_path,info)


workbook.save(path_output)

with open(speakers_file,"w+") as file:
    for i,speaker in enumerate(speaker_names):
        file.write(f"{i} {speaker}\n")

print(f"Данные добавлены в файл {path_output} ")

        






    